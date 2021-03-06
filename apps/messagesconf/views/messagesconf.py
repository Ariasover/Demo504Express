""" User's Views """

# Django 
from django.contrib.auth import authenticate,logout,login
from django.views.generic import View
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render,get_object_or_404
from django.conf import settings
from django.urls import reverse
from django.views.generic import ListView,UpdateView,CreateView,DeleteView
from django.views.generic.edit import FormView
from django.contrib import messages
from django.db import transaction


# Selenium
from selenium.webdriver.common.by import By
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities


# Models
from apps.messagesconf.models.messagesconf import *


# Utils
from apps.utils.chrome_version import get_chrome_version


# Forms
from ..forms import *
from ...speech.forms import MessagesConfigurationForm


# Openpyxl
from openpyxl import load_workbook


# Local
import socket,os,plistlib,datetime
from pathlib import Path
from sys import platform
from time import sleep
import time as tiempo
import re


class ConfigurationView(View):
	def get(self, request):
		return render(request, 'configuration.html')


class DashboardAereoView(ListView):
	model = MessagesList
	template_name = 'dashboard_aereo.html'
	paginate_by = 5
	context_object_name = "not_sent_messages"

	# Status
	no_enviado = MessageListStatus.objects.get(description__icontains='No Enviado')
	enviado = MessageListStatus.objects.get(description='Enviado')
	error = MessageListStatus.objects.get(description__icontains='Error')
	
	
	def verify_excel(self, departure_date):
		message_list = MessagesList.objects.filter(departure_date=departure_date)
		for instance in message_list:
    		# Verify all about phone
			if not len(instance.phone) == 11:
				phone = re.sub('[\.-]','', instance.phone)
				message_list.filter(pk=instance.pk).update(phone = phone)
		messages.success(self.request, '¡Archivo subido y verificado con éxito!', extra_tags='success')


	def send_messages(self):
		# Configuration
		status_not_sent = MessageListStatus.objects.get(description__icontains='No Enviado')
		customer_list = MessagesList.objects.filter(status=status_not_sent).values('pk','name','phone', 'message','amount','departure_date','weight_greather','weight_type')		
		message_configuration = MessagesConfiguration.objects.get(is_active=True).text
		message_text=""
		xpath = '//*[@id="main"]/footer/div[1]/div[2]/div/div[2]'
		invalid_xpath = '/html/body/div[1]/div/span[2]/div/span/div/div/div/div/div/div[2]/div/div/div'
		time=20

		
		if customer_list.exists():	 
			if platform == "linux" or platform == "linux2" or platform == "darwin":
				plistloc = "/Applications/Google Chrome.app/Contents/Info.plist"
				pl = plistlib.readPlist(plistloc)
				chrome_server_version = pl["CFBundleShortVersionString"]
				chrome_server_version = chrome_server_version[0]+chrome_server_version[1] #for example '87' #TODO CAMBIAR ESTE METODO POR EL DE CHROMDRIVER()PARA OBTENER MEJOR LA VERSION
				driver = webdriver.Chrome(executable_path=str(settings.VIRTUALENV_DIR)+'/lib/python3.7/site-packages/chromedriver_autoinstaller/'+chrome_server_version+'/chromedriver')
			else:
				
				driver = webdriver.Remote(
   					command_executor='http://172.31.29.22:4445/wd/hub',
   					desired_capabilities=DesiredCapabilities.CHROME)
				
			driver.get("http://web.whatsapp.com")
			# sleep(5) # Cambiar si es necesario
			text_box=""
			# startTime = tiempo.time()
			for count,customer in enumerate(customer_list):
				try:
					message_text = message_configuration.replace('/name/',customer['name'])
					message_text = message_text.replace('/fecha/',str(customer['departure_date']))
					message_text = message_text.replace('/monto/','{:0,.2f}'.format(customer['amount']))
					message_text = message_text.replace('/pesomayor/',customer['weight_greather'])
					message_text = message_text.replace('/tipo_peso/',customer['weight_type'])
					driver.get("https://web.whatsapp.com/send?phone={}&source=&data=#".format(str(customer['phone'])))
					
					try:
						driver.switch_to_alert().accept()
					except Exception as e:
						print('Error',e)
										
					invalid_number_popup = driver.find_elements_by_xpath(invalid_xpath)
					if not invalid_number_popup:
						WebDriverWait(driver, time).until(EC.presence_of_element_located((By.XPATH, xpath)))
						text_box=driver.find_element(By.XPATH, xpath)
						for part in message_text.split('\n'):
							text_box.send_keys(part)
							ActionChains(driver).key_down(Keys.SHIFT).key_down(Keys.ENTER).key_up(Keys.SHIFT).key_up(Keys.ENTER).perform()
						ActionChains(driver).send_keys(Keys.RETURN).perform()

						customer_list.filter(pk=customer['pk']).update(status = self.enviado)					
					else:	
						error = ErrorNumber()
						error.message_list = MessagesList.objects.get(pk=customer['pk'])
						error.creation_user=self.request.user
						error.modification_user=self.request.user
						error.save()

				except Exception as e:
					print('NOT SENT ===',e)
					if not ErrorNumber.objects.filter(message_list=customer['pk']).exists():	
						error = ErrorNumber()
						error.message_list = MessagesList.objects.get(pk=customer['pk'])
						error.creation_user=self.request.user
						error.modification_user=self.request.user
						error.save()
				print('mensaje',count+1)
				if count == 15:
					break

			# endTime = tiempo.time()
			# elapsedTime = endTime - startTime
			# print("Elapsed Time = %s" % elapsedTime)
			messages.success(self.request, 'Mensajes enviados',extra_tags='success')
			driver.quit()
		else:
			messages.error(self.request, 'No hay destinatarios',extra_tags='error')


	def upload_excel(self,request):
		try: 
			with transaction.atomic():
				# Verify extension
				file = self.request.FILES['myfile']
				if file.name.endswith('xlsx'):
						
					doc = load_workbook(file,data_only=True)
					nombres_hojas = doc.sheetnames
					hoja1 = doc.get_sheet_by_name(nombres_hojas[0])	
					departure_date = str(hoja1['M2'].value)

					if MessagesList.objects.filter(departure_date = departure_date).exists():
						messages.error(self.request, 'El archivo ya existe',extra_tags='error')
					else:

						# Initial configuration						
						mayor = 0     						
						# Get total number of rows
						for y in range(4,1000):
							if not hoja1['B'+str(y)].value == None:
								mayor=mayor+1
							else:
								break
						
						# Save in database
						for i in range (4, mayor+4):	
							if not hoja1['N'+str(i)].value == '#VALUE!' or not hoja1['N'+str(i)].value == '#N/A':
								if  not hoja1['C'+str(i)].value == None:
									messages_list = MessagesList()
									messages_list.name=hoja1['B'+str(i)].value
									messages_list.phone = "504"+str(hoja1['C'+str(i)].value)
									messages_list.departure_date = departure_date
									messages_list.amount = hoja1['P'+str(i)].value
									messages_list.weight_greather = hoja1['N'+str(i)].value
									messages_list.weight_type = hoja1['M'+str(i)].value
									messages_list.creation_user=self.request.user
									messages_list.modification_user=self.request.user
									messages_list.status = MessageListStatus.objects.get(description__icontains='No Enviado')
									messages_list.save()

						# Verify if excel is valid
						self.verify_excel(departure_date)
				else:
					messages.error(request, 'La extensión del archivo es incorrecta', extra_tags='error')
		except Exception as e:
			print('error',e)
			messages.error(request, 'No se ha podido subir el archivo', extra_tags='error')


	def get_queryset(self):
		"""Returns number of not sent items"""
		not_sent = MessagesList.objects.filter()
		return not_sent
	

	def post(self, request):
		if request.POST['options'] == "send_all":
			self.send_messages()
		elif request.POST['options'] == 'upload_excel':
			self.upload_excel(request)
		return HttpResponseRedirect(reverse('messagesconf:dashboard_aereo'))
	

	def get_context_data(self, **kwargs):
		context = super().get_context_data(**kwargs)
		status_not_sent = MessageListStatus.objects.get(description__icontains='No Enviado')
		not_sent= self.get_queryset().filter(status=status_not_sent).count()
		total = self.get_queryset().count()
		sent = self.get_queryset().filter(status=self.enviado).count()
		speech = MessagesConfiguration.objects.filter(is_active=True).get()
		context.update({
			'not_sent':not_sent,
			'sent':sent,
			'total':total,
			'speech':speech,
		})		
		return context
		
			
class OneMessageView(FormView):
	model = MessagesList 
	template_name = 'one_message.html'
	form_class = OneMessageForm

	def send_message(self,phone,message):	
		message_text=message
		xpath = '//*[@id="main"]/footer/div[1]/div[2]/div/div[2]'
		invalid_xpath = '/html/body/div[1]/div/span[2]/div/span/div/div/div/div/div/div[2]/div/div/div'
		time=20
		try:
			if platform == "linux" or platform == "linux2" or platform == "darwin":
				plistloc = "/Applications/Google Chrome.app/Contents/Info.plist"
				pl = plistlib.readPlist(plistloc)
				chrome_server_version = pl["CFBundleShortVersionString"]
				chrome_server_version = chrome_server_version[0]+chrome_server_version[1] #for example '87' #TODO CAMBIAR ESTE METODO POR EL DE CHROMDRIVER()PARA OBTENER MEJOR LA VERSION
				driver = webdriver.Chrome(executable_path=str(settings.VIRTUALENV_DIR)+'/lib/python3.7/site-packages/chromedriver_autoinstaller/'+chrome_server_version+'/chromedriver')
			else:
				driver = webdriver.Remote(
					command_executor='http://localhost:4444/wd/hub',
					desired_capabilities=DesiredCapabilities.CHROME)
				
			driver.get("http://web.whatsapp.com")
			# sleep(5) # Cambiar si es necesario
			text_box=""


			driver.get("https://web.whatsapp.com/send?phone={}&source=&data=#".format(str(phone)))
				# TODO
			try:
				print('MAMADO2')
				driver.switch_to_alert().accept()
				print('MAMADO3')
			except Exception as e:
				print('Error',e)					
			sleep(5)

			invalid_number_popup = driver.find_elements_by_xpath(invalid_xpath)
			if not invalid_number_popup:
				WebDriverWait(driver, time).until(EC.presence_of_element_located((By.XPATH, xpath)))
				text_box=driver.find_element(By.XPATH, xpath)
				message_text = message_text.split('\n')
				for part in message_text:
					text_box.send_keys(part)
					ActionChains(driver).key_down(Keys.SHIFT).key_down(Keys.ENTER).key_up(Keys.SHIFT).key_up(Keys.ENTER).perform()
				ActionChains(driver).send_keys(Keys.RETURN).perform()
				messages.success(self.request, 'Mensaje enviado',extra_tags='success')
				# customer_list.filter(pk=customer.pk).update(status = self.enviado) #Todo
			else:	
				messages.error(self.request, 'Mensaje no enviado',extra_tags='error')

		except Exception as e:
			print('NOT SENT ===',e)
			messages.error(self.request, 'Mensaje no enviado',extra_tags='error')
			# if not ErrorNumber.objects.filter(message_list=customer).exists():	
			# 	error = ErrorNumber()
			# 	error.message_list = customer
			# 	error.creation_user=self.request.user
			# 	error.modification_user=self.request.user
			# 	error.save()
		sleep(3)
		
		driver.quit()


	def get_success_url(self):
		return reverse('messagesconf:one_message')


	def form_valid(self, form):
		self.object = form.save(commit=False)
		self.object.creation_user = self.request.user
		self.object.modification_user = self.request.user
		self.object.status = MessageListStatus.objects.get(description='Enviado')
		self.object.amount = 0
		self.object.departure_date = datetime.datetime.now()
		self.object.weight_greather = 0
		self.object.weight_type = 0
		self.object.save()
		phone = self.object.phone
		message = self.object.message

		self.send_message(phone,message)

		return HttpResponseRedirect(self.get_success_url())